import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
from io import BytesIO
import time
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, date, timedelta

# ── Page config ──────────────────────────────────────────────
st.set_page_config(page_title="SEC DR Filing Monitor", page_icon="📋", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Thai:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans Thai', sans-serif; }
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

.app-header {
    background: linear-gradient(135deg, #1a3a5c 0%, #0d2137 100%);
    border-radius: 12px; padding: 24px 32px; margin-bottom: 20px;
    border-left: 5px solid #3498db;
}
.app-header h1 { color: white; margin: 0; font-size: 1.7rem; font-weight: 600; }
.app-header p  { color: #a8c4de; margin: 5px 0 0; font-size: 0.9rem; }

.metric-row { display: flex; gap: 12px; margin-bottom: 18px; flex-wrap: wrap; }
.metric-card {
    flex: 1; min-width: 130px; background: white;
    border-radius: 10px; padding: 16px 18px; border-top: 4px solid;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
.metric-card .num { font-size: 2rem; font-weight: 700; line-height: 1; font-family: 'IBM Plex Mono', monospace; }
.metric-card .lbl { font-size: 0.75rem; color: #666; margin-top: 4px; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
.card-total  { border-color: #1a3a5c; } .card-total .num  { color: #1a3a5c; }
.card-s1     { border-color: #95a5a6; } .card-s1 .num     { color: #7f8c8d; }
.card-s2     { border-color: #3498db; } .card-s2 .num     { color: #2980b9; }
.card-s3     { border-color: #27ae60; } .card-s3 .num     { color: #27ae60; }

.box {
    background: #f8f9fa; border-radius: 10px;
    padding: 16px 20px; margin-bottom: 16px; border: 1px solid #e9ecef;
}
.box-title { font-weight: 600; font-size: 0.88rem; color: #1a3a5c;
             text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 12px; }
.email-box { background: #fff8e1; border-radius: 10px;
             padding: 16px 20px; border: 1px solid #ffe082; margin-bottom: 16px; }
</style>
""", unsafe_allow_html=True)

# ── Constants ────────────────────────────────────────────────
ISSUER_CODES = {
    "บัวหลวง": "01", "bualuang": "01", "bls": "01",
    "พาย": "03", "pi": "03", "pai": "03",
    "เกียรตินาคิน": "06", "ภัทร": "06", "kiatnakin": "06", "phatra": "06", "kkp": "06",
    "เคเจไอ": "13", "kgi": "13",
    "หยวนต้า": "19", "yuanta": "19",
    "อินโนเวสท์": "23", "innovest": "23", "scbx": "23",
    "กรุงไทย": "80", "ktb": "80", "ktbst": "80", "ธนาคารกรุงไทย": "80",
}
HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "th-TH,th;q=0.9,en;q=0.8",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}
DEFAULT_RECIPIENTS = "nuttaphat.huayudomsin@krungthai.com"


# ── Helpers ──────────────────────────────────────────────────
def get_issuer_code(name):
    lower = (name or "").lower()
    for key, code in ISSUER_CODES.items():
        if key.lower() in lower:
            return code
    return ""

def generate_set_symbol(underlying, issuer):
    if not underlying or underlying == "—":
        return ""
    code = get_issuer_code(issuer)
    if not code:
        return ""
    ticker = re.sub(r"\d+$", "", underlying.strip().split()[0].upper())
    return f"{ticker}{code}"

def parse_th_date(s):
    """DD/MM/YYYY (BE or CE) → Python date, or None."""
    if not s or str(s).strip() in ("", "—"):
        return None
    try:
        d, m, y = str(s).strip().split("/")
        y = int(y)
        if y > 2400:
            y -= 543
        return date(y, int(m), int(d))
    except Exception:
        return None

def detect_stage(first, amend, effective, trade_start):
    if trade_start and str(trade_start).strip():
        return "3. Trading Started"
    if effective and str(effective).strip():
        return "2. Filing Effective"
    if first and str(first).strip():
        return "1. Initial Filing"
    return "—"

def scrape_underlying(detail_url):
    try:
        r = requests.get(detail_url, headers=HEADERS, timeout=10)
        r.encoding = "TIS-620"
        text = BeautifulSoup(r.text, "html.parser").get_text(" ")
        idx = text.find("ผู้เสนอขายหลักทรัพย์")
        if idx == -1:
            return ""
        segment = text[idx: idx + 400]
        for p in reversed(re.findall(r"\(([^)]{3,})\)", segment)):
            if re.search(r"[A-Za-z]{3,}", p):
                return p.strip()
    except Exception:
        pass
    return ""

def parse_filings_html(html):
    """
    SEC ViewMore table columns (0-indexed):
      0  ผู้ออกหลักทรัพย์/ผู้เสนอขายหลักทรัพย์   → issuer
      1  ประเภทหลักทรัพย์                          → sec_type
      2  ประเภทการเสนอขาย                          → offer_type
      3  วันที่ยื่น Filing version แรก             → first_date
      4  วันที่แก้ไขข้อมูล Filing ล่าสุด           → amend_date
      5  วันที่ Filing มีผลบังคับ                  → effective
      6  วันที่เริ่มการเสนอขาย                     → trade_start
      7  วันที่สิ้นสุดการขาย                       → offer_end
      8  หมายเหตุ                                  → remark
      9  Filing (icon link)                         → detail_url
    """
    filings = []
    seen = set()  # dedup by (issuer, first_date)

    for row in BeautifulSoup(html, "html.parser").find_all("tr"):
        if row.find("th"):
            continue
        cells = row.find_all("td")
        if len(cells) < 7:
            continue

        def cell(i):
            return cells[i].get_text(strip=True) if i < len(cells) else ""

        # Extract issuer — take part before "/" (e.g. "บัวหลวง / ธนาคารกรุงไทย" → "บัวหลวง")
        issuer_raw = cells[0].get_text(" ", strip=True)
        issuer     = issuer_raw.split("/")[0].strip()

        first_date  = cell(3)
        amend_date  = cell(4)
        effective   = cell(5)
        trade_start = cell(6)
        offer_end   = cell(7)
        remark      = cell(8)

        # Detail URL: find <a> tags with href containing "capital.sec.or.th" or "final69"
        # Fall back to last <a> in the row
        detail_url = ""
        for a in cells[-1].find_all("a", href=True):
            detail_url = a["href"]
            break
        if not detail_url:
            for a in row.find_all("a", href=True):
                href = a["href"]
                if "capital.sec.or.th" in href or "final69" in href or "cgi-bin" in href:
                    detail_url = href
                    break
            if not detail_url:
                all_links = row.find_all("a", href=True)
                detail_url = all_links[-1]["href"] if all_links else ""

        # Dedup: skip exact duplicate (issuer + first_date)
        dedup_key = (issuer.strip(), first_date.strip())
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        filings.append({
            "issuer":      issuer,
            "sec_type":    cell(1),
            "offer_type":  cell(2),
            "first_date":  first_date,
            "amend_date":  amend_date,
            "effective":   effective,
            "trade_start": trade_start,
            "offer_end":   offer_end,
            "remark":      remark,
            "detail_url":  detail_url,
            "underlying":  "",   # filled later by enrich step
            "set_symbol":  "",
            "set_link":    "",
            "stage":       "",
        })
    return filings

def fetch_and_enrich(date_from, date_to, progress_bar, status_text):
    all_raw  = []
    seen_global = set()   # cross-page dedup

    for page in range(34):
        url = f"https://market.sec.or.th/public/idisc/th/ViewMore/filing-equity?SecuTypeCode=DS&FilingData={page}"
        status_text.text(f"📄 Scanning page {page + 1}/34...")
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.encoding = "UTF-8"
            page_filings = parse_filings_html(r.text)  # already per-page deduped
            if not page_filings:
                status_text.text(f"✅ No more data at page {page + 1}.")
                break

            oldest = None
            added  = 0
            for f in page_filings:
                d = parse_th_date(f["first_date"])
                if d:
                    if oldest is None or d < oldest:
                        oldest = d
                    if date_from <= d <= date_to:
                        key = (f["issuer"].strip(), f["first_date"].strip())
                        if key not in seen_global:
                            seen_global.add(key)
                            all_raw.append(f)
                            added += 1

            progress_bar.progress(min((page + 1) / 34 * 0.4, 0.4))
            status_text.text(f"📄 Page {page + 1}: +{added} in range (total so far: {len(all_raw)})")

            if oldest and oldest < date_from:
                status_text.text(f"✅ Oldest filing {oldest} < {date_from} — stopping scan.")
                break
            time.sleep(0.5)
        except Exception as e:
            status_text.text(f"⚠️ Page {page + 1} error: {e}")
            break

    total = len(all_raw)
    status_text.text(f"🔍 Found {total} unique filings in range. Fetching underlying names...")

    for i, f in enumerate(all_raw):
        status_text.text(f"🔍 [{i+1}/{total}] Fetching detail: {f['issuer'][:40]}...")
        progress_bar.progress(0.4 + (i + 1) / max(total, 1) * 0.6)

        if f.get("detail_url"):
            underlying = scrape_underlying(f["detail_url"])
            f["underlying"] = underlying if underlying else "—"
        else:
            f["underlying"] = "—"

        sym = generate_set_symbol(f["underlying"], f["issuer"])
        f["set_symbol"] = sym
        f["set_link"]   = f"https://www.set.or.th/en/market/product/dr/quote/{sym}/price" if sym else ""
        f["stage"]      = detect_stage(f["first_date"], f["amend_date"], f["effective"], f["trade_start"])

        if i < total - 1:
            time.sleep(0.5)

    progress_bar.progress(1.0)
    status_text.text(f"✅ Done! {total} unique filings fetched.")
    return all_raw

def to_dataframe(filings):
    rows = []
    for f in filings:
        # Guard: underlying must be a string company name, not a stage label
        underlying = str(f.get("underlying") or "—").strip()
        if underlying.startswith("1.") or underlying.startswith("2.") or underlying.startswith("3."):
            underlying = "—"

        set_sym = str(f.get("set_symbol") or "").strip()
        # Guard: set symbol must look like TICKER## not a date
        if not re.match(r'^[A-Z]{2,}[0-9]{2,}$', set_sym):
            set_sym = "—"

        rows.append({
            "Underlying Stock":         underlying,
            "SET Symbol":               set_sym or "—",
            "Stage":                    str(f.get("stage") or "—"),
            "Issuer":                   str(f.get("issuer") or ""),
            "Offer Type":               str(f.get("offer_type") or ""),
            "วันที่ยื่น Filing แรก":    str(f.get("first_date") or ""),
            "วันที่แก้ไขล่าสุด":        str(f.get("amend_date") or ""),
            "วันที่ Filing มีผลบังคับ": str(f.get("effective") or ""),
            "วันที่เริ่มเทรด":          str(f.get("trade_start") or ""),
            "วันที่สิ้นสุดการขาย":      str(f.get("offer_end") or ""),
            "Security Type":            str(f.get("sec_type") or ""),
            "หมายเหตุ":                 str(f.get("remark") or ""),
            "SEC Filing URL":           str(f.get("detail_url") or ""),
            "SET Link":                 str(f.get("set_link") or ""),
        })
    return pd.DataFrame(rows)

def to_excel(df):
    out = BytesIO()
    # Reorder for Excel output
    excel_cols = [
        "Underlying Stock", "SET Symbol", "Stage", "Issuer", "Offer Type",
        "วันที่ยื่น Filing แรก", "วันที่แก้ไขล่าสุด", "วันที่ Filing มีผลบังคับ",
        "วันที่เริ่มเทรด", "วันที่สิ้นสุดการขาย", "Security Type", "หมายเหตุ",
        "SEC Filing URL", "SET Link",
    ]
    export_df = df[[c for c in excel_cols if c in df.columns]]
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="DR Filings")
        ws = writer.sheets["DR Filings"]
        from openpyxl.styles import PatternFill, Font, Alignment
        col_widths = [32, 14, 22, 35, 16, 20, 20, 24, 20, 22, 20, 15, 55, 45]
        for i, w in enumerate(col_widths, 1):
            if i <= ws.max_column:
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        hfill = PatternFill("solid", fgColor="1a3a5c")
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 40
        ws.freeze_panes = "A2"
        stage_colors = {
            "3. Trading Started":  "C8E6C9",
            "2. Filing Effective": "BBDEFB",
            "1. Initial Filing":   "FFF9C4",
        }
        sc = next((c.column for c in ws[1] if c.value == "Stage"), None)
        if sc:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                color = stage_colors.get(str(row[sc - 1].value or ""))
                if color:
                    fill = PatternFill("solid", fgColor=color)
                    for cell in row:
                        cell.fill = fill
    out.seek(0)
    return out


# ── Email builders ───────────────────────────────────────────
STAGE_COLOR = {
    "3. Trading Started":  "#27ae60",
    "2. Filing Effective": "#2980b9",
    "1. Initial Filing":   "#7f8c8d",
}

def _row_html(r):
    sym      = r["SET Symbol"] if r["SET Symbol"] != "—" else ""
    set_link = r.get("SET Link", "")
    sec_link = r.get("SEC Filing URL", "")
    sc       = STAGE_COLOR.get(r["Stage"], "#ccc")
    return f"""<tr>
      <td style="padding:9px 12px;border:1px solid #ddd;font-weight:bold;color:#1a5276;">{r["Underlying Stock"]}</td>
      <td style="padding:9px 12px;border:1px solid #ddd;">
        <span style="background:{sc};color:white;padding:3px 10px;border-radius:12px;font-size:11px;">{r["Stage"]}</span>
      </td>
      <td style="padding:9px 12px;border:1px solid #ddd;">{r["Issuer"]}</td>
      <td style="padding:9px 12px;border:1px solid #ddd;">{r["วันที่ยื่น Filing แรก"]}</td>
      <td style="padding:9px 12px;border:1px solid #ddd;color:#27ae60;font-weight:bold;">{r["วันที่เริ่มเทรด"] or "—"}</td>
      <td style="padding:9px 12px;border:1px solid #ddd;font-weight:bold;">
        {'<a href="' + set_link + '" style="color:#c0392b;text-decoration:none;">' + sym + '</a>' if sym else '—'}
      </td>
      <td style="padding:9px 12px;border:1px solid #ddd;text-align:center;">
        {'<a href="' + sec_link + '" style="background:#2980b9;color:white;padding:4px 10px;border-radius:4px;text-decoration:none;font-size:12px;">SEC →</a>' if sec_link else '—'}
      </td>
    </tr>"""

TABLE_HEAD = """<tr style="background:#2c3e50;color:white;">
  <th style="padding:10px;text-align:left;">Underlying</th>
  <th style="padding:10px;text-align:left;">Stage</th>
  <th style="padding:10px;text-align:left;">Issuer</th>
  <th style="padding:10px;text-align:left;">วันที่ยื่น Filing แรก</th>
  <th style="padding:10px;text-align:left;">วันที่เริ่มเทรด</th>
  <th style="padding:10px;text-align:left;">SET Symbol</th>
  <th style="padding:10px;text-align:left;">Filing</th>
</tr>"""

def build_weekly_html(df, date_from, date_to):
    rows = "".join(_row_html(r) for _, r in df.iterrows()) or \
           '<tr><td colspan="7" style="padding:14px;text-align:center;color:#888;">No filings in this range</td></tr>'
    return f"""<div style="font-family:Arial,sans-serif;max-width:1100px;">
      <div style="background:#1a5276;color:white;padding:16px 24px;border-radius:6px 6px 0 0;">
        <h2 style="margin:0;">🆕 DR Filing Weekly Report</h2>
        <p style="margin:6px 0 0;font-size:13px;opacity:0.85;">
          {date_from.strftime("%d/%m/%Y")} – {date_to.strftime("%d/%m/%Y")} &nbsp;·&nbsp; <b>{len(df)} รายการ</b>
        </p>
      </div>
      <table style="border-collapse:collapse;width:100%;font-size:13px;border:1px solid #ddd;">
        <thead>{TABLE_HEAD}</thead><tbody>{rows}</tbody>
      </table>
      <p style="color:#aaa;font-size:11px;margin-top:20px;">SEC DR Filing Monitor · {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>
    </div>"""

def build_monthly_html(df):
    def section(title, color, sub):
        if sub.empty:
            return f"<h3 style='color:{color}'>{title} (0)</h3><p style='color:#888;font-size:13px;'>No records</p>"
        rows = "".join(_row_html(r) for _, r in sub.iterrows())
        return f"""<h3 style="margin:24px 0 8px;color:{color}">{title} ({len(sub)})</h3>
        <table style="border-collapse:collapse;width:100%;font-size:13px;border:1px solid #ddd;">
          <thead>{TABLE_HEAD}</thead><tbody>{rows}</tbody>
        </table>"""

    s1 = df[df["Stage"] == "1. Initial Filing"]
    s2 = df[df["Stage"] == "2. Filing Effective"]
    s3 = df[df["Stage"] == "3. Trading Started"]
    return f"""<div style="font-family:Arial,sans-serif;max-width:1100px;">
      <div style="background:#1a5276;color:white;padding:20px 24px;border-radius:6px 6px 0 0;">
        <h2 style="margin:0;">📋 Monthly DR Filing Summary — {datetime.now().strftime("%B %Y")}</h2>
      </div>
      <div style="background:#eaf4fb;padding:16px 24px;border:1px solid #d6eaf8;display:flex;gap:40px;">
        <div style="text-align:center;"><div style="font-size:28px;font-weight:bold;color:#1a5276;">{len(df)}</div><div style="font-size:12px;color:#555;">Total</div></div>
        <div style="text-align:center;"><div style="font-size:28px;font-weight:bold;color:#7f8c8d;">{len(s1)}</div><div style="font-size:12px;color:#555;">Initial Filing</div></div>
        <div style="text-align:center;"><div style="font-size:28px;font-weight:bold;color:#2980b9;">{len(s2)}</div><div style="font-size:12px;color:#555;">Filing Effective</div></div>
        <div style="text-align:center;"><div style="font-size:28px;font-weight:bold;color:#27ae60;">{len(s3)}</div><div style="font-size:12px;color:#555;">Trading Started</div></div>
      </div>
      {section("1. Initial Filing", "#7f8c8d", s1)}
      {section("2. Filing Effective", "#2980b9", s2)}
      {section("3. Trading Started", "#27ae60", s3)}
      <p style="color:#aaa;font-size:11px;margin-top:20px;">SEC DR Filing Monitor · {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>
    </div>"""

def send_email(smtp_user, smtp_pass, recipients, subject, html_body):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(recipients)
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(smtp_user, smtp_pass)
        s.sendmail(smtp_user, recipients, msg.as_string())


# ══════════════════════════════════════════════════════════════
# LAYOUT
# ══════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
  <h1>📋 SEC DR Filing Monitor</h1>
  <p>ตราสารแสดงสิทธิในหลักทรัพย์ต่างประเทศ (DR) · Real-time data from SEC Thailand</p>
</div>
""", unsafe_allow_html=True)

tab_data, tab_email = st.tabs(["📊  Data & Fetch", "✉️  Email Settings & Send"])


# ══════════════════════════════════════════════════════════════
# TAB 1 — DATA
# ══════════════════════════════════════════════════════════════
with tab_data:

    # ── Date range controls ──────────────────────────────────
    st.markdown('<div class="box">', unsafe_allow_html=True)
    st.markdown('<div class="box-title">📅 Date Range & Fetch</div>', unsafe_allow_html=True)

    r1c1, r1c2, r1c3, r1c4 = st.columns([2, 2, 1, 2])
    with r1c1:
        date_from = st.date_input("From (วันที่ยื่น Filing แรก)",
                                  value=date.today() - timedelta(days=30),
                                  max_value=date.today())
    with r1c2:
        date_to = st.date_input("To", value=date.today(), max_value=date.today())
    with r1c3:
        st.markdown("<br>", unsafe_allow_html=True)
        fetch_btn = st.button("🔄 Fetch", type="primary", use_container_width=True)
    with r1c4:
        preset = st.selectbox("Quick preset", ["— custom —", "Last 7 days", "Last 30 days",
                                                "Last 90 days", "This year"],
                              label_visibility="collapsed")

    # Apply preset (overrides pickers)
    _preset_map = {
        "Last 7 days":  (date.today() - timedelta(days=7),  date.today()),
        "Last 30 days": (date.today() - timedelta(days=30), date.today()),
        "Last 90 days": (date.today() - timedelta(days=90), date.today()),
        "This year":    (date(date.today().year, 1, 1),      date.today()),
    }
    if preset in _preset_map:
        date_from, date_to = _preset_map[preset]
        st.caption(f"Using preset: **{preset}** → {date_from} to {date_to}")

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Run fetch ────────────────────────────────────────────
    if fetch_btn:
        if date_from > date_to:
            st.error("❌ 'From' date must be before 'To' date.")
        else:
            pb   = st.progress(0)
            stxt = st.empty()
            filings = fetch_and_enrich(date_from, date_to, pb, stxt)
            st.session_state.update({
                "filings":    filings,
                "df":         to_dataframe(filings),
                "fetched_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "date_from":  date_from,
                "date_to":    date_to,
            })
            time.sleep(0.3)
            st.rerun()

    # ── Show results ─────────────────────────────────────────
    if "df" in st.session_state and not st.session_state["df"].empty:
        df         = st.session_state["df"]
        fetched_at = st.session_state.get("fetched_at", "")
        d_from     = st.session_state.get("date_from", "")
        d_to       = st.session_state.get("date_to", "")

        # Metric cards
        s1c = len(df[df["Stage"] == "1. Initial Filing"])
        s2c = len(df[df["Stage"] == "2. Filing Effective"])
        s3c = len(df[df["Stage"] == "3. Trading Started"])
        st.markdown(f"""
        <div class="metric-row">
          <div class="metric-card card-total"><div class="num">{len(df)}</div><div class="lbl">Total Filings</div></div>
          <div class="metric-card card-s1"><div class="num">{s1c}</div><div class="lbl">1 · Initial Filing</div></div>
          <div class="metric-card card-s2"><div class="num">{s2c}</div><div class="lbl">2 · Filing Effective</div></div>
          <div class="metric-card card-s3"><div class="num">{s3c}</div><div class="lbl">3 · Trading Started</div></div>
        </div>""", unsafe_allow_html=True)

        # Filters
        st.markdown('<div class="box">', unsafe_allow_html=True)
        st.markdown('<div class="box-title">🔎 Filters</div>', unsafe_allow_html=True)
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            stage_f  = st.multiselect("Stage",
                ["1. Initial Filing", "2. Filing Effective", "3. Trading Started"])
        with fc2:
            issuer_f = st.multiselect("Issuer",
                sorted(df["Issuer"].dropna().unique()))
        with fc3:
            offer_f  = st.multiselect("Offer Type",
                sorted(df["Offer Type"].dropna().unique()))
        with fc4:
            search   = st.text_input("🔍 Search", placeholder="Underlying / Symbol / Issuer…")
        st.markdown('</div>', unsafe_allow_html=True)

        filt = df.copy()
        if stage_f:  filt = filt[filt["Stage"].isin(stage_f)]
        if issuer_f: filt = filt[filt["Issuer"].isin(issuer_f)]
        if offer_f:  filt = filt[filt["Offer Type"].isin(offer_f)]
        if search:
            m = (filt["Underlying Stock"].str.contains(search, case=False, na=False) |
                 filt["SET Symbol"].str.contains(search, case=False, na=False) |
                 filt["Issuer"].str.contains(search, case=False, na=False))
            filt = filt[m]

        st.caption(f"**{len(filt)} filings shown** · Fetched: {fetched_at} · Range: {d_from} → {d_to}")

        DISPLAY = [
            "Underlying Stock", "SET Symbol", "Stage", "Issuer", "Offer Type",
            "วันที่ยื่น Filing แรก", "วันที่แก้ไขล่าสุด",
            "วันที่ Filing มีผลบังคับ", "วันที่เริ่มเทรด", "SEC Filing URL"
        ]

        # Use explicit hex colors that work on both light & dark Streamlit themes
        STAGE_BG = {
            "3. Trading Started":  "#1e5c36",  # dark green text-readable
            "2. Filing Effective": "#1a3a5c",  # dark blue
            "1. Initial Filing":   "#5c4a00",  # dark amber
        }
        STAGE_FG = {
            "3. Trading Started":  "#a8f0c0",
            "2. Filing Effective": "#a8d0f0",
            "1. Initial Filing":   "#f0d870",
        }

        def style_stage(val):
            bg = STAGE_BG.get(val, "")
            fg = STAGE_FG.get(val, "")
            if bg:
                return f"background-color:{bg};color:{fg};font-weight:600;border-radius:4px;padding:2px 6px;"
            return ""

        display_df = filt[DISPLAY].copy()
        # Make SEC Filing URL a shorter label for display
        display_df = display_df.rename(columns={"SEC Filing URL": "SEC Link"})

        st.dataframe(
            display_df.style.applymap(style_stage, subset=["Stage"]),
            use_container_width=True, height=500, hide_index=True,
            column_config={
                "SEC Link": st.column_config.LinkColumn("SEC Link", display_text="View →"),
                "SET Symbol": st.column_config.TextColumn("SET Symbol", width="small"),
                "Stage": st.column_config.TextColumn("Stage", width="medium"),
            }
        )

        st.markdown("---")
        dc1, dc2, _ = st.columns([2, 2, 4])
        with dc1:
            st.download_button("📥 Download Filtered Excel", data=to_excel(filt),
                file_name=f"DR_filtered_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with dc2:
            st.download_button("📥 Download ALL Excel", data=to_excel(df),
                file_name=f"DR_ALL_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    else:
        st.info("👆 Set a date range and press **Fetch** to load data.")
        st.markdown("""
**How it works:**
- Scans SEC filing pages newest → oldest, stops automatically once it passes your start date
- Fetches each detail page to get the underlying stock name (TIS-620 encoded)
- 💡 Shorter ranges = faster — last 30 days typically takes 2–5 min
        """)


# ══════════════════════════════════════════════════════════════
# TAB 2 — EMAIL
# ══════════════════════════════════════════════════════════════
with tab_email:

    # ── SMTP credentials ─────────────────────────────────────
    st.markdown('<div class="box">', unsafe_allow_html=True)
    st.markdown('<div class="box-title">⚙️ Gmail SMTP Credentials</div>', unsafe_allow_html=True)
    st.caption("Stored only in your browser session — never saved anywhere.")

    ec1, ec2 = st.columns(2)
    with ec1:
        smtp_user = st.text_input("Gmail address (sender)",
            value=st.session_state.get("smtp_user", ""),
            placeholder="your.email@gmail.com")
    with ec2:
        smtp_pass = st.text_input("Gmail App Password", type="password",
            value=st.session_state.get("smtp_pass", ""),
            placeholder="xxxx xxxx xxxx xxxx",
            help="myaccount.google.com → Security → 2-Step Verification → App passwords")

    if smtp_user: st.session_state["smtp_user"] = smtp_user
    if smtp_pass: st.session_state["smtp_pass"] = smtp_pass
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Recipients ───────────────────────────────────────────
    st.markdown('<div class="box">', unsafe_allow_html=True)
    st.markdown('<div class="box-title">📬 Recipients</div>', unsafe_allow_html=True)
    st.caption("One email per line. Edits are kept for this session.")

    if "recipients_text" not in st.session_state:
        st.session_state["recipients_text"] = DEFAULT_RECIPIENTS

    rec_text = st.text_area("Recipients", value=st.session_state["recipients_text"],
                             height=110, label_visibility="collapsed",
                             placeholder="email1@example.com\nemail2@example.com")
    st.session_state["recipients_text"] = rec_text

    recipients = [r.strip() for r in rec_text.splitlines() if r.strip() and "@" in r]
    if recipients:
        st.success(f"✅ {len(recipients)} recipient(s): {', '.join(recipients)}")
    else:
        st.warning("⚠️ No valid recipients entered.")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Send buttons ─────────────────────────────────────────
    st.markdown('<div class="email-box">', unsafe_allow_html=True)
    st.markdown('<div class="box-title">📤 Send Email</div>', unsafe_allow_html=True)

    has_data = "df" in st.session_state and not st.session_state["df"].empty
    if not has_data:
        st.warning("⚠️ No data loaded — go to the **Data & Fetch** tab and fetch first.")

    sb1, sb2 = st.columns(2)
    with sb1:
        st.markdown("**📅 Weekly Email**")
        st.caption("Current fetched data as a concise weekly report.")
        send_weekly = st.button("📤 Send Weekly Email",
                                disabled=not has_data, use_container_width=True)
    with sb2:
        st.markdown("**📋 Monthly Summary**")
        st.caption("Full summary grouped by stage (Initial / Effective / Trading).")
        send_monthly = st.button("📤 Send Monthly Email",
                                 disabled=not has_data, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    def _check():
        if not st.session_state.get("smtp_user") or not st.session_state.get("smtp_pass"):
            st.error("❌ Enter Gmail address and App Password above.")
            return False
        if not recipients:
            st.error("❌ Add at least one recipient.")
            return False
        return True

    if send_weekly and _check():
        with st.spinner("Sending weekly email…"):
            try:
                df      = st.session_state["df"]
                d_from  = st.session_state.get("date_from", date.today() - timedelta(days=7))
                d_to    = st.session_state.get("date_to",   date.today())
                html    = build_weekly_html(df, d_from, d_to)
                subject = (f"🆕 [SEC DR] Weekly Report "
                           f"{d_from.strftime('%d/%m/%Y')}–{d_to.strftime('%d/%m/%Y')} "
                           f"· {len(df)} filings")
                send_email(st.session_state["smtp_user"], st.session_state["smtp_pass"],
                           recipients, subject, html)
                st.success(f"✅ Weekly email sent to: {', '.join(recipients)}")
            except Exception as e:
                st.error(f"❌ Failed: {e}")

    if send_monthly and _check():
        with st.spinner("Sending monthly summary…"):
            try:
                df      = st.session_state["df"]
                html    = build_monthly_html(df)
                subject = (f"📋 [SEC DR] Monthly Summary {datetime.now().strftime('%B %Y')} "
                           f"· {len(df)} filings")
                send_email(st.session_state["smtp_user"], st.session_state["smtp_pass"],
                           recipients, subject, html)
                st.success(f"✅ Monthly email sent to: {', '.join(recipients)}")
            except Exception as e:
                st.error(f"❌ Failed: {e}")

    # ── Help ─────────────────────────────────────────────────
    with st.expander("ℹ️ How to create a Gmail App Password"):
        st.markdown("""
1. Go to [myaccount.google.com](https://myaccount.google.com)
2. **Security → 2-Step Verification** (must be enabled)
3. Scroll down → **App passwords**
4. App: **Mail** · Device: **Other** → name it `SEC Monitor`
5. Copy the 16-character password → paste in the field above

> ⚠️ Treat this like a password — don't share it.
        """)
